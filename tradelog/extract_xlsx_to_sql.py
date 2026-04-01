"""
Extract data from 'Portfolio-Tracking 2026.xlsx' and generate a SQL Server script
to populate the tradelog database.

Usage:
    python tradelog/extract_xlsx_to_sql.py [--ibkr-account-id 8497] [--output tradelog/seed.sql]

The generated .sql file can be run manually against the database via:
    sqlcmd -S localhost -d tradelog -i tradelog/seed.sql
or via SSMS / Azure Data Studio.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / "Portfolio-Tracking 2026.xlsx"

# ── Enum mappings (Excel display value → DB int) ────────────────────────

BUDGET = {"Core": 0, "Speculative": 1}

STRATEGY = {
    "Positive Drift": 0,
    "Range Bound Commodities": 1,
    "PEADS": 2,
    "Momentum": 3,
    "IV mean reversion": 4,
    "IV Mean Reversion": 4,
    "Sector strength": 5,
    "Sector Strength": 5,
    "Sector weakness": 6,
    "Sector Weakness": 6,
    "Breakout": 7,
    "Green Line Breakout": 8,
    "Slingshot": 9,
    "Pre-Earnings": 10,
    "PreEarnings": 10,
    "Pre Earnings": 10,
    "Pre earnings": 10,
}

TYPE_OF_TRADE = {
    "Short Strangle": 0,
    "Short Put Spread": 1,
    "Short Call Spread": 2,
    "Long Call": 3,
    "Long Put": 4,
    "Long Call Vertical": 5,
    "Long Put Vertical": 6,
    "Syntetic Long": 7,  # typo in spreadsheet
    "Synthetic Long": 7,
    "Covered Strangle": 8,
    "Butterfly": 9,
    "Ratio Diagonal Spread": 10,
    "Long Strangle": 11,
    "Short Put": 12,
    "Short Call": 13,
    "Long Stock": 14,
    "Short Stock": 15,
    "Short Put Vertical": 1,  # alias for Short Put Spread
}

DIRECTIONAL = {
    "Bullish": 0,
    "Neutral": 1,
    "Bearish": 2,
    # Multi-value entries — map to primary bias
    "Neutral, Bullish": 0,
    "Bullish, Neutral": 0,
    "Neutral, Bearish": 2,
    "Bullish, Bearish": 1,  # mixed → Neutral
}

TIMEFRAME = {"1d": 0, "1w": 1, "Delta Band": 2}

MANAGEMENT_RATING = {
    "As planned": 0,
    "As Planned": 0,
    "Minor adjustments": 1,
    "Minor Adjustments": 1,
    "Rogue": 2,
    "Rouge": 2,  # typo in spreadsheet
}

POSITION_RIGHT = {"C": 0, "Call": 0, "P": 1, "Put": 1}


# ── Helpers ──────────────────────────────────────────────────────────────


def sql_str(v: str | None) -> str:
    """Escape a string for T-SQL, returning NULL for None/empty."""
    if v is None:
        return "NULL"
    s = str(v).strip()
    if not s:
        return "NULL"
    return "N'" + s.replace("'", "''") + "'"


def sql_datetime(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%dT%H:%M:%S')}'"
    if isinstance(v, (int, float)):
        # Excel serial date
        base = datetime(1899, 12, 30)
        return f"'{(base + timedelta(days=v)).strftime('%Y-%m-%dT%H:%M:%S')}'"
    return "NULL"


def sql_num(v, default: str = "NULL") -> str:
    if v is None:
        return default
    if isinstance(v, float) and (v != v or v == float("inf") or v == float("-inf")):
        return default  # NaN or Inf
    if isinstance(v, str):
        if v.startswith("="):
            return default  # formula, skip
        if v.lower().strip() in ("nan", "inf", "-inf", ""):
            return default
        try:
            return str(float(v))
        except ValueError:
            return default
    return str(v)


def sql_int(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, str):
        if v.startswith("="):
            return "NULL"
        try:
            return str(int(float(v)))
        except ValueError:
            return "NULL"
    return str(int(v))


def sql_bool(v) -> str:
    if v is None:
        return "0"
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, str):
        return "1" if v.lower() in ("true", "yes", "1") else "0"
    return "1" if v else "0"


def sql_enum(v, mapping: dict, default: str = "NULL") -> str:
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    result = mapping.get(s)
    if result is None:
        # Try case-insensitive
        for k, val in mapping.items():
            if k.lower() == s.lower():
                return str(val)
        print(f"  WARNING: unknown enum value '{s}' — mapped to {default}", file=sys.stderr)
        return default
    return str(result)


def is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


# ── Extractors ───────────────────────────────────────────────────────────


def extract_portfolios(wb, account_id: int | str) -> list[str]:
    ws = wb["Portfolio"]
    stmts = []
    for row in ws.iter_rows(min_row=2):
        budget_raw = row[0].value
        strategy_raw = row[1].value
        if budget_raw is None or strategy_raw is None:
            continue
        budget = sql_enum(budget_raw, BUDGET)
        strategy = sql_enum(strategy_raw, STRATEGY)
        if budget == "NULL" or strategy == "NULL":
            continue
        min_alloc = sql_num(row[2].value)
        max_alloc = sql_num(row[3].value)
        stmts.append(
            f"INSERT INTO Portfolios (AccountId, Budget, Strategy, MinAllocation, MaxAllocation) "
            f"VALUES ({account_id}, {budget}, {strategy}, {min_alloc}, {max_alloc});"
        )
    return stmts


def extract_weekly_preps(wb, account_id: int | str) -> list[str]:
    ws = wb["WeeklyPrep"]
    stmts = []
    for row in ws.iter_rows(min_row=2):
        date_val = row[0].value
        if date_val is None:
            continue
        stmts.append(
            f"INSERT INTO WeeklyPreps (AccountId, Date, IndexBias, Breadth, NotableSectors, "
            f"VolatilityNotes, OpenPositionsRequiringManagement, CurrentPortfolioRisk, "
            f"PortfolioNotes, ScanningFor, IndexSectorPreference, Watchlist, Learnings, "
            f"FocusForImprovement, ExternalComments) VALUES ("
            f"{account_id}, {sql_datetime(date_val)}, "
            f"{sql_str(row[1].value)}, {sql_str(row[2].value)}, {sql_str(row[3].value)}, "
            f"{sql_str(row[4].value)}, {sql_str(row[5].value)}, {sql_str(row[6].value)}, "
            f"{sql_str(row[7].value)}, {sql_str(row[8].value)}, {sql_str(row[9].value)}, "
            f"{sql_str(row[10].value)}, {sql_str(row[11].value)}, "
            f"{sql_str(row[12].value)}, {sql_str(row[13].value)});"
        )
    return stmts


def extract_capitals(wb, account_id: int | str) -> list[str]:
    """Extract only the manually entered columns from Capital sheet."""
    ws = wb["Capital"]
    stmts = []
    for row in ws.iter_rows(min_row=2):
        date_val = row[0].value
        if date_val is None:
            continue
        net_liq = row[1].value
        maint = row[2].value
        excess = row[3].value
        bpr = row[4].value
        # Skip if all manual fields are empty
        if all(v is None for v in [net_liq, maint, excess, bpr]):
            continue
        # MaintenancePct = Maintenance * 100 / NetLiquidity
        maint_pct = 0
        if net_liq and maint and float(net_liq) != 0:
            maint_pct = round(float(maint) * 100 / float(net_liq), 6)

        stmts.append(
            f"INSERT INTO Capitals (AccountId, Date, NetLiquidity, Maintenance, "
            f"ExcessLiquidity, Bpr, MaintenancePct, TotalPnl, UnrealizedPnl, "
            f"RealizedPnl, NetDelta, NetTheta, NetVega, NetGamma, AvgIv, "
            f"TotalMargin, TotalCommissions) VALUES ("
            f"{account_id}, {sql_datetime(date_val)}, "
            f"{sql_num(net_liq)}, {sql_num(maint)}, {sql_num(excess)}, {sql_num(bpr)}, "
            f"{maint_pct}, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0);"
        )
    return stmts


def extract_trade_entries(wb, account_id: int | str) -> list[str]:
    """Extract from TradeLog sheet → TradeEntries table."""
    ws = wb["TradeLog"]
    stmts = []
    # Headers: Symbol(0), Date(1), TypeOfTrade(2), Notes(3), Directional(4),
    # Timeframe(5), Budget(6), Strategy(7), NewsCatalyst(8), RecentEarnings(9),
    # SectorSupport(10), ATH(11), RVOL(12), Put/Call(13), xATRmove(14),
    # InstitutionalSupport(15), Gap%(16), TA/FA(17), IntendedManagement(18),
    # Management(19), ManagementRating(20), Learnings(21)
    for row in ws.iter_rows(min_row=2):
        symbol = row[0].value
        date_val = row[1].value
        if symbol is None or date_val is None:
            continue
        if is_formula(symbol):
            continue

        stmts.append(
            f"INSERT INTO TradeEntries (AccountId, Symbol, Date, TypeOfTrade, Notes, "
            f"Directional, Timeframe, Budget, Strategy, NewsCatalyst, RecentEarnings, "
            f"SectorSupport, Ath, Rvol, XAtrMove, InstitutionalSupport, GapPct, "
            f"TaFaNotes, IntendedManagement, ActualManagement, ManagementRating, "
            f"Learnings) VALUES ("
            f"{account_id}, {sql_str(symbol)}, {sql_datetime(date_val)}, "
            f"{sql_enum(row[2].value, TYPE_OF_TRADE, '0')}, "
            f"{sql_str(row[3].value)}, "
            f"{sql_enum(row[4].value, DIRECTIONAL, '0')}, "
            f"{sql_enum(row[5].value, TIMEFRAME, '0')}, "
            f"{sql_enum(row[6].value, BUDGET, '0')}, "
            f"{sql_enum(row[7].value, STRATEGY, '0')}, "
            f"{sql_bool(row[8].value)}, {sql_bool(row[9].value)}, "
            f"{sql_bool(row[10].value)}, {sql_bool(row[11].value)}, "
            f"{sql_num(row[12].value)}, "
            f"{sql_num(row[14].value)}, "  # xATR move (col 14)
            f"{sql_str(row[15].value)}, "  # Institutional support
            f"{sql_num(row[16].value)}, "  # Gap%
            f"{sql_str(row[17].value)}, "  # TA/FA
            f"{sql_str(row[18].value)}, "  # Intended management
            f"{sql_str(row[19].value)}, "  # Actual management
            f"{sql_enum(row[20].value, MANAGEMENT_RATING)}, "
            f"{sql_str(row[21].value)});"
        )
    return stmts


def extract_option_positions(wb, account_id: int | str) -> list[str]:
    """Extract raw data from OptionPositions sheet."""
    ws = wb["OptionPositions"]
    stmts = []
    # Headers: Symbol(0), ContractId(1), Opened(2), Expiry(3), Closed(4),
    # Pos(5), Right(6), Strike(7), Cost(8), Commission(9), Multiplier(10)
    for row in ws.iter_rows(min_row=2):
        symbol = row[0].value
        contract_id = row[1].value
        if symbol is None or contract_id is None:
            continue
        if is_formula(symbol):
            continue

        opened = row[2].value
        expiry = row[3].value
        closed = row[4].value
        pos = row[5].value
        right_val = row[6].value
        strike = row[7].value
        cost = row[8].value
        commission = row[9].value
        multiplier = row[10].value

        stmts.append(
            f"INSERT INTO OptionPositions (AccountId, Symbol, ContractId, Opened, Expiry, "
            f"Closed, Pos, [Right], Strike, Cost, Commission, Multiplier) VALUES ("
            f"{account_id}, {sql_str(symbol)}, {sql_str(str(contract_id))}, "
            f"{sql_datetime(opened)}, {sql_datetime(expiry)}, {sql_datetime(closed)}, "
            f"{sql_int(pos)}, {sql_enum(right_val, POSITION_RIGHT, '0')}, "
            f"{sql_num(strike)}, {sql_num(cost)}, {sql_num(commission)}, "
            f"{sql_int(multiplier) if multiplier else '100'});"
        )
    return stmts


def extract_stock_trades(wb, account_id: int | str) -> list[str]:
    """Extract from StockTrades sheet → Trades table."""
    ws = wb["StockTrades"]
    stmts = []
    # Headers: Symbol(0), Date(1), PosChange(2), Price(3), ..., Commission(7)
    for row in ws.iter_rows(min_row=2):
        symbol = row[0].value
        date_val = row[1].value
        if symbol is None or date_val is None:
            continue
        if is_formula(symbol):
            continue

        pos_change = row[2].value
        price = row[3].value
        commission = row[7].value

        stmts.append(
            f"INSERT INTO Trades (AccountId, Symbol, Date, PosChange, Price, Commission, Multiplier) "
            f"VALUES ({account_id}, {sql_str(symbol)}, {sql_datetime(date_val)}, "
            f"{sql_int(pos_change)}, {sql_num(price)}, {sql_num(commission, '0')}, 1);"
        )
    return stmts


def extract_option_positions_log(wb, account_id: int | str) -> list[str]:
    """Extract from OptionPositionsLog sheet."""
    ws = wb["OptionPositionsLog"]
    stmts = []
    # Headers: DateTime(0), ContractId(1), Underlying(2), IV(3), Price(4),
    # TimeValue(5), Delta(6), Theta(7), Gamma(8), Vega(9), Margin(10)
    for row in ws.iter_rows(min_row=2):
        dt = row[0].value
        contract_id = row[1].value
        if dt is None or contract_id is None:
            continue

        stmts.append(
            f"INSERT INTO OptionPositionsLogs (AccountId, DateTime, ContractId, "
            f"Underlying, Iv, Price, TimeValue, Delta, Theta, Gamma, Vega, Margin) "
            f"VALUES ({account_id}, {sql_datetime(dt)}, {sql_str(str(contract_id))}, "
            f"{sql_num(row[2].value, '0')}, {sql_num(row[3].value, '0')}, "
            f"{sql_num(row[4].value, '0')}, {sql_num(row[5].value, '0')}, "
            f"{sql_num(row[6].value, '0')}, {sql_num(row[7].value, '0')}, "
            f"{sql_num(row[8].value, '0')}, {sql_num(row[9].value, '0')}, "
            f"{sql_num(row[10].value, '0')});"
        )
    return stmts


# ── Main ─────────────────────────────────────────────────────────────────


def emit_account_preamble(lines: list[str], ibkr_account_id: str) -> None:
    """Shared SQL preamble: resolve account by IbkrAccountId or abort."""
    lines.append("SET NOCOUNT ON;")
    lines.append("BEGIN TRANSACTION;")
    lines.append("")
    lines.append("DECLARE @AccountId INT;")
    lines.append(
        f"SELECT @AccountId = Id FROM Accounts "
        f"WHERE IbkrAccountId = N'{ibkr_account_id}';"
    )
    lines.append("")
    lines.append("IF @AccountId IS NULL")
    lines.append("BEGIN")
    lines.append(
        f"    PRINT 'ERROR: No account found for IbkrAccountId {ibkr_account_id}';")
    lines.append("    ROLLBACK;")
    lines.append("    RETURN;")
    lines.append("END")
    lines.append("")
    lines.append(
        "PRINT 'Using AccountId ' + CAST(@AccountId AS VARCHAR) + "
        f"' (IBKR: {ibkr_account_id})';"
    )
    lines.append("")


def run_full_seed(wb, args) -> None:
    """Original mode: wipe account data and re-seed everything from Excel."""
    output_path = args.output or str(Path(args.xlsx).with_suffix(".sql"))
    aid = "@AccountId"
    lines: list[str] = []

    lines.append("-- ==========================================================")
    lines.append("-- Auto-generated FULL seed script from Portfolio-Tracking 2026.xlsx")
    lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"-- IBKR Account: {args.ibkr_account_id}")
    lines.append("-- ==========================================================")
    lines.append("")

    # Full seed creates account if missing
    lines.append("SET NOCOUNT ON;")
    lines.append("BEGIN TRANSACTION;")
    lines.append("")
    lines.append("DECLARE @AccountId INT;")
    lines.append(
        f"SELECT @AccountId = Id FROM Accounts "
        f"WHERE IbkrAccountId = N'{args.ibkr_account_id}';"
    )
    lines.append("")
    lines.append("IF @AccountId IS NULL")
    lines.append("BEGIN")
    lines.append(
        f"    INSERT INTO Accounts (IbkrAccountId, Name, Host, Port, ClientId, IsDefault) "
        f"VALUES (N'{args.ibkr_account_id}', N'IBKR {args.ibkr_account_id}', "
        f"N'127.0.0.1', 7497, 1, 1);"
    )
    lines.append("    SET @AccountId = SCOPE_IDENTITY();")
    lines.append(
        f"    PRINT 'Created account for IbkrAccountId {args.ibkr_account_id} "
        f"with Id ' + CAST(@AccountId AS VARCHAR);"
    )
    lines.append("END")
    lines.append("")
    lines.append("PRINT 'Using AccountId ' + CAST(@AccountId AS VARCHAR) + "
                 f"' (IBKR: {args.ibkr_account_id})';")
    lines.append("")

    lines.append("-- Remove all existing data for this account")
    for table in [
        "OptionPositionsLogs",
        "OptionPositions",
        "Trades",
        "TradeEntries",
        "Capitals",
        "WeeklyPreps",
        "Portfolios",
    ]:
        lines.append(f"DELETE FROM {table} WHERE AccountId = @AccountId;")
    lines.append("")

    sections = [
        ("Portfolios", extract_portfolios(wb, aid)),
        ("WeeklyPreps", extract_weekly_preps(wb, aid)),
        ("Capitals", extract_capitals(wb, aid)),
        ("TradeEntries", extract_trade_entries(wb, aid)),
        ("OptionPositions", extract_option_positions(wb, aid)),
        ("Trades", extract_stock_trades(wb, aid)),
        ("OptionPositionsLogs", extract_option_positions_log(wb, aid)),
    ]

    for table_name, stmts in sections:
        if not stmts:
            lines.append(f"-- {table_name}: no data found")
            lines.append("")
            continue
        lines.append(f"-- {table_name} ({len(stmts)} rows)")
        lines.append(f"PRINT 'Inserting {len(stmts)} rows into {table_name}...';")
        lines.extend(stmts)
        lines.append("")

    lines.append("COMMIT TRANSACTION;")
    lines.append("PRINT 'Seed complete.';")

    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    print(f"Written {len(lines)} lines to {output_path}")
    for table_name, stmts in sections:
        print(f"  {table_name}: {len(stmts)} rows")


def extract_options_history_positions(wb, account_id: int | str) -> list[str]:
    """Generate INSERT-if-not-exists statements for OptionPositions."""
    ws = wb["OptionPositions"]
    stmts = []
    for row in ws.iter_rows(min_row=2):
        symbol = row[0].value
        contract_id = row[1].value
        if symbol is None or contract_id is None:
            continue
        if is_formula(symbol):
            continue

        cid = sql_str(str(contract_id))
        stmts.append(
            f"IF NOT EXISTS (SELECT 1 FROM OptionPositions "
            f"WHERE AccountId = {account_id} AND ContractId = {cid})\n"
            f"  INSERT INTO OptionPositions (AccountId, Symbol, ContractId, ConId, "
            f"Opened, Expiry, Closed, Pos, [Right], Strike, Cost, Commission, Multiplier) "
            f"VALUES ({account_id}, {sql_str(symbol)}, {cid}, "
            f"{sql_int(contract_id)}, "
            f"{sql_datetime(row[2].value)}, {sql_datetime(row[3].value)}, "
            f"{sql_datetime(row[4].value)}, "
            f"{sql_int(row[5].value)}, {sql_enum(row[6].value, POSITION_RIGHT, '0')}, "
            f"{sql_num(row[7].value)}, {sql_num(row[8].value)}, "
            f"{sql_num(row[9].value)}, "
            f"{sql_int(row[10].value) if row[10].value else '100'});"
        )
    return stmts


def extract_options_history_logs(wb, account_id: int | str) -> list[str]:
    """Generate INSERT-if-not-exists statements for OptionPositionsLog."""
    ws = wb["OptionPositionsLog"]
    stmts = []
    for row in ws.iter_rows(min_row=2):
        dt = row[0].value
        contract_id = row[1].value
        if dt is None or contract_id is None:
            continue

        cid = sql_str(str(contract_id))
        dt_sql = sql_datetime(dt)
        stmts.append(
            f"IF NOT EXISTS (SELECT 1 FROM OptionPositionsLogs "
            f"WHERE AccountId = {account_id} AND ContractId = {cid} AND DateTime = {dt_sql})\n"
            f"  INSERT INTO OptionPositionsLogs (AccountId, DateTime, ContractId, "
            f"Underlying, Iv, Price, TimeValue, Delta, Theta, Gamma, Vega, Margin) "
            f"VALUES ({account_id}, {dt_sql}, {cid}, "
            f"{sql_num(row[2].value, '0')}, {sql_num(row[3].value, '0')}, "
            f"{sql_num(row[4].value, '0')}, {sql_num(row[5].value, '0')}, "
            f"{sql_num(row[6].value, '0')}, {sql_num(row[7].value, '0')}, "
            f"{sql_num(row[8].value, '0')}, {sql_num(row[9].value, '0')}, "
            f"{sql_num(row[10].value, '0')});"
        )
    return stmts


def run_options_history(wb, args) -> None:
    """Import only OptionPositions + OptionPositionsLog, skipping existing rows."""
    output_path = args.output or str(
        Path(args.xlsx).with_name("import-options-history.sql")
    )
    aid = "@AccountId"
    lines: list[str] = []

    lines.append("-- ==========================================================")
    lines.append("-- Import historical option positions + Greeks logs (skip existing)")
    lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"-- IBKR Account: {args.ibkr_account_id}")
    lines.append("-- ==========================================================")
    lines.append("")

    emit_account_preamble(lines, args.ibkr_account_id)

    sections = [
        ("OptionPositions", extract_options_history_positions(wb, aid)),
        ("OptionPositionsLogs", extract_options_history_logs(wb, aid)),
    ]

    for table_name, stmts in sections:
        if not stmts:
            lines.append(f"-- {table_name}: no data found")
            lines.append("")
            continue
        lines.append(f"-- {table_name} ({len(stmts)} rows, skip existing)")
        lines.append(f"PRINT 'Importing up to {len(stmts)} rows into {table_name}...';")
        lines.extend(stmts)
        lines.append("")

    lines.append("COMMIT TRANSACTION;")
    lines.append("PRINT 'Options history import complete.';")

    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    print(f"Written {len(lines)} lines to {output_path}")
    for table_name, stmts in sections:
        print(f"  {table_name}: {len(stmts)} rows")


def main():
    parser = argparse.ArgumentParser(description="Extract XLSX → SQL seed script")
    parser.add_argument(
        "--ibkr-account-id",
        type=str,
        default="8497",
        help="IBKR Account ID to link data to (default: 8497)",
    )
    parser.add_argument("--output", type=str, default=None, help="Output .sql file path")
    parser.add_argument("--xlsx", type=str, default=str(XLSX_PATH), help="Input xlsx path")
    parser.add_argument(
        "--mode",
        choices=["full", "options-history"],
        default="full",
        help="full = wipe & reseed all data (default), "
             "options-history = import only OptionPositions + logs (skip existing)",
    )
    args = parser.parse_args()

    print(f"Reading {args.xlsx} (mode={args.mode}) ...")
    wb = openpyxl.load_workbook(args.xlsx, data_only=False)

    if args.mode == "options-history":
        run_options_history(wb, args)
    else:
        run_full_seed(wb, args)


if __name__ == "__main__":
    main()
