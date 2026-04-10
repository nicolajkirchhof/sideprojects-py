"""
One-off merge importer for the Tradelog stage database.

Reads one or more Portfolio-Tracking xlsx files (2025 + 2026 layouts supported),
inserts missing rows idempotently, and interactively links newly-imported
trades to existing OptionPositions / StockPositions by symbol + date (±1 day).

Usage:
    uv run --with openpyxl --with pyodbc python tradelog/import_from_xlsx.py \
        --server ocin.database.windows.net \
        --database tradelog_stage \
        --user ocin \
        --password <pwd> \
        --ibkr-account-id U16408041 \
        --xlsx "tradelog/Portfolio-Tracking 2025.xlsx" \
        --xlsx "tradelog/Portfolio-Tracking 2026.xlsx"

Add --non-interactive to skip position-link prompts and log unmatched trades.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import pyodbc


# ── Enum mappings (copied/extended from extract_xlsx_to_sql.py) ────────
# DB int values match the C# enums in the Models project.

BUDGET = {"Core": 0, "Speculative": 1}

STRATEGY = {
    "Positive Drift": 0, "Range Bound Commodities": 1, "PEADS": 2, "Momentum": 3,
    "IV mean reversion": 4, "IV Mean Reversion": 4,
    "Sector strength": 5, "Sector Strength": 5,
    "Sector weakness": 6, "Sector Weakness": 6,
    "Breakout": 7, "Green Line Breakout": 8, "Slingshot": 9,
    "Pre-Earnings": 10, "PreEarnings": 10, "Pre Earnings": 10, "Pre earnings": 10,
}

TYPE_OF_TRADE = {
    "Short Strangle": 0, "SS": 0, "STR": 0,
    "Short Put Spread": 1, "SPS": 1, "Short Put Vertical": 1,
    "Short Call Spread": 2, "SCS": 2, "Short Call Vertical": 2,
    "Long Call": 3, "LC": 3,
    "Long Put": 4, "LP": 4,
    "Long Call Vertical": 5, "LCV": 5, "Long Call Spread": 5,
    "Long Put Vertical": 6, "LPV": 6, "Long Put Spread": 6,
    "Syntetic Long": 7, "Synthetic Long": 7, "SL": 7,
    "Covered Strangle": 8, "CS": 8,
    "Butterfly": 9, "BF": 9,
    "Ratio Diagonal Spread": 10, "RDS": 10,
    "Long Strangle": 11, "LS": 11,
    "Short Put": 12, "SP": 12, "NP": 12,
    "Short Call": 13, "SC": 13, "NC": 13,
    "Long Stock": 14,
    "Short Stock": 15,
    "Jade Lizard": 0, "Jade Lizzard": 0,  # variant of short strangle
    "Paid Call": 3,  # funded long call
    "Inverted Strangle": 0, "Short Straddle": 0,
    "Double Calendar": 9,  # similar to butterfly
    "PDS": 1, "Put Debit": 6, "Put Debit Spread": 6,
}

DIRECTIONAL = {
    "Bullish": 0, "Neutral": 1, "Bearish": 2,
    "Neutral, Bullish": 0, "Bullish, Neutral": 0,
    "Neutral, Bearish": 2, "Bearish, Neutral": 2,
    "Bullish, Bearish": 1,
}

TIMEFRAME = {"1d": 0, "1w": 1, "Delta Band": 2, "1h": 0}

MANAGEMENT_RATING = {
    "As planned": 0, "As Planned": 0,
    "Minor adjustments": 1, "Minor Adjustments": 1,
    "Rogue": 2, "Rouge": 2,
}

POSITION_RIGHT = {"C": 0, "Call": 0, "P": 1, "Put": 1}


# ── Helpers ─────────────────────────────────────────────────────────────

def clean_val(v: Any) -> Any:
    """Normalize Excel cell values — drop formulas, NaN, empty strings."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("=") or s == "#REF!":
            return None
        return s
    if isinstance(v, float):
        if v != v:  # NaN
            return None
    return v


def enum_val(raw: Any, mapping: dict, default: int | None = None) -> int | None:
    v = clean_val(raw)
    if v is None:
        return default
    s = str(v).strip()
    if s in mapping:
        return mapping[s]
    for k, val in mapping.items():
        if k.lower() == s.lower():
            return val
    print(f"  WARN unknown enum: {s!r}", file=sys.stderr)
    return default


def num_val(raw: Any) -> float | None:
    v = clean_val(raw)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def int_val(raw: Any) -> int | None:
    v = clean_val(raw)
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def bool_val(raw: Any) -> bool:
    v = clean_val(raw)
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.lower() in ("true", "yes", "1")
    return bool(v)


def str_val(raw: Any) -> str | None:
    v = clean_val(raw)
    return None if v is None else str(v)


def safe_decimal(raw: Any) -> float:
    """Returns a safe float for SQL decimal columns — 0.0 for any non-parseable value."""
    v = num_val(raw)
    if v is None or v != v or v == float("inf") or v == float("-inf"):
        return 0.0
    return float(v)


def date_val(raw: Any) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(raw))
    return None


# ── DB helpers ─────────────────────────────────────────────────────────

def connect(server: str, database: str, user: str, password: str) -> pyodbc.Connection:
    # Pick the best installed ODBC driver for SQL Server (18 preferred, 17 fallback).
    available = pyodbc.drivers()
    driver = next(
        (d for d in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server")
         if d in available),
        None,
    )
    if driver is None:
        raise RuntimeError(f"No suitable SQL Server ODBC driver found. Installed: {available}")
    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={server};DATABASE={database};UID={user};PWD={password};"
        "Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=120;"
    )
    return pyodbc.connect(conn_str, autocommit=False)


def resolve_account_id(cursor: pyodbc.Cursor, ibkr_id: str) -> int:
    cursor.execute("SELECT Id FROM Accounts WHERE IbkrAccountId = ?", ibkr_id)
    row = cursor.fetchone()
    if row is None:
        raise RuntimeError(f"No account found for IbkrAccountId={ibkr_id}")
    return int(row[0])


# ── Row extractors ─────────────────────────────────────────────────────

@dataclass
class TradeRow:
    symbol: str
    date: datetime
    type_of_trade: int | None
    notes: str | None
    directional: int | None
    timeframe: int | None
    budget: int
    strategy: int | None
    news_catalyst: bool
    recent_earnings: bool
    sector_support: bool
    ath: bool
    rvol: float | None
    institutional_support: str | None
    gap_pct: float | None
    x_atr_move: float | None
    ta_fa: str | None
    intended_management: str | None
    actual_management: str | None
    management_rating: int | None
    learnings: str | None


def extract_trades_2026(ws) -> list[TradeRow]:
    rows: list[TradeRow] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        def g(i: int) -> Any:
            return r[i] if i < len(r) else None
        symbol = str_val(g(0))
        dt = date_val(g(1))
        if not symbol or not dt:
            continue
        rows.append(TradeRow(
            symbol=symbol.upper(),
            date=dt,
            type_of_trade=enum_val(g(2), TYPE_OF_TRADE),
            notes=str_val(g(3)),
            directional=enum_val(g(4), DIRECTIONAL),
            timeframe=enum_val(g(5), TIMEFRAME),
            budget=enum_val(g(6), BUDGET, default=0) or 0,
            strategy=enum_val(g(7), STRATEGY),
            news_catalyst=bool_val(g(8)),
            recent_earnings=bool_val(g(9)),
            sector_support=bool_val(g(10)),
            ath=bool_val(g(11)),
            rvol=num_val(g(12)),
            institutional_support=str_val(g(15)),
            gap_pct=num_val(g(16)),
            x_atr_move=num_val(g(14)),
            ta_fa=str_val(g(17)),
            intended_management=str_val(g(18)),
            actual_management=str_val(g(19)),
            management_rating=enum_val(g(20), MANAGEMENT_RATING),
            learnings=str_val(g(21)),
        ))
    return rows


def extract_trades_2025(ws) -> list[TradeRow]:
    # 2025 headers: Symbol(0), Date(1), TypeOfTrade(2), Notes(3), Sentiment(4),
    # Directional(5), Timeframe(6), Strategy(7), NewsCatalyst(8), RecentEarnings(9),
    # SectorSupport(10), TA/FA(11), StopLoss(12), Management(13),
    # ManagementRating(14), Learnings(15)
    # Rows may be shorter than 16 cells if trailing cells are empty.
    rows: list[TradeRow] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        def g(i: int) -> Any:
            return r[i] if i < len(r) else None
        symbol = str_val(g(0))
        dt = date_val(g(1))
        if not symbol or not dt:
            continue
        rows.append(TradeRow(
            symbol=symbol.upper(),
            date=dt,
            type_of_trade=enum_val(g(2), TYPE_OF_TRADE),
            notes=str_val(g(3)),
            directional=enum_val(g(5), DIRECTIONAL),
            timeframe=enum_val(g(6), TIMEFRAME),
            budget=0,  # 2025 file has no Budget column → default to Core
            strategy=enum_val(g(7), STRATEGY),
            news_catalyst=bool_val(g(8)),
            recent_earnings=bool_val(g(9)),
            sector_support=bool_val(g(10)),
            ath=False,
            rvol=None,
            institutional_support=None,
            gap_pct=None,
            x_atr_move=None,
            ta_fa=str_val(g(11)),
            intended_management=None,
            actual_management=str_val(g(13)),
            management_rating=enum_val(g(14), MANAGEMENT_RATING),
            learnings=str_val(g(15)),
        ))
    return rows


# ── Inserters ──────────────────────────────────────────────────────────

def insert_trade_if_missing(cursor: pyodbc.Cursor, account_id: int, t: TradeRow) -> int | None:
    """Returns the inserted Trade.Id, or None if the row already existed."""
    cursor.execute(
        """
        SELECT TOP 1 Id FROM Trades
        WHERE AccountId = ? AND Symbol = ? AND Date = ?
          AND (Strategy = ? OR (? IS NULL AND Strategy IS NULL))
        """,
        account_id, t.symbol, t.date, t.strategy, t.strategy,
    )
    existing = cursor.fetchone()
    if existing:
        return None

    # Strategy and TypeOfTrade are NOT NULL in the DB — default to 0 when unknown.
    strategy = t.strategy if t.strategy is not None else 0
    type_of_trade = t.type_of_trade if t.type_of_trade is not None else 0

    cursor.execute(
        """
        INSERT INTO Trades
            (AccountId, Symbol, Date, TypeOfTrade, Notes, Directional, Timeframe,
             Budget, Strategy, NewsCatalyst, RecentEarnings, SectorSupport, Ath,
             Rvol, InstitutionalSupport, GapPct, XAtrMove, TaFaNotes,
             IntendedManagement, ActualManagement, ManagementRating, Learnings)
        OUTPUT INSERTED.Id
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        account_id, t.symbol, t.date, type_of_trade, t.notes, t.directional,
        t.timeframe, t.budget, strategy, t.news_catalyst, t.recent_earnings,
        t.sector_support, t.ath, t.rvol, t.institutional_support, t.gap_pct,
        t.x_atr_move, t.ta_fa, t.intended_management, t.actual_management,
        t.management_rating, t.learnings,
    )
    return int(cursor.fetchone()[0])


def insert_weeklyprep_if_missing(cursor: pyodbc.Cursor, account_id: int, row: tuple) -> int:
    """Returns 1 if inserted, 0 if already existed."""
    dt = date_val(row[0] if len(row) > 0 else None)
    if dt is None:
        return 0
    cursor.execute("SELECT 1 FROM WeeklyPreps WHERE AccountId = ? AND Date = ?", account_id, dt)
    if cursor.fetchone():
        return 0
    cursor.execute(
        """
        INSERT INTO WeeklyPreps
            (AccountId, Date, IndexBias, Breadth, NotableSectors, VolatilityNotes,
             OpenPositionsRequiringManagement, CurrentPortfolioRisk, PortfolioNotes,
             ScanningFor, IndexSectorPreference, Watchlist, Learnings,
             FocusForImprovement, ExternalComments)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        account_id, dt,
        str_val(row[1] if len(row) > 1 else None),
        str_val(row[2] if len(row) > 2 else None),
        str_val(row[3] if len(row) > 3 else None),
        str_val(row[4] if len(row) > 4 else None),
        str_val(row[5] if len(row) > 5 else None),
        str_val(row[6] if len(row) > 6 else None),
        str_val(row[7] if len(row) > 7 else None),
        str_val(row[8] if len(row) > 8 else None),
        str_val(row[9] if len(row) > 9 else None),
        str_val(row[10] if len(row) > 10 else None),
        str_val(row[11] if len(row) > 11 else None),
        str_val(row[12] if len(row) > 12 else None),
        str_val(row[13] if len(row) > 13 else None),
    )
    return 1


def insert_capital_if_missing(cursor: pyodbc.Cursor, account_id: int, row: tuple) -> int:
    dt = date_val(row[0] if len(row) > 0 else None)
    if dt is None:
        return 0
    net_liq = num_val(row[1] if len(row) > 1 else None)
    maint = num_val(row[2] if len(row) > 2 else None)
    excess = num_val(row[3] if len(row) > 3 else None)
    bpr = num_val(row[4] if len(row) > 4 else None)
    if net_liq is None and maint is None and excess is None and bpr is None:
        return 0
    cursor.execute("SELECT 1 FROM Capitals WHERE AccountId = ? AND Date = ?", account_id, dt)
    if cursor.fetchone():
        return 0
    maint_pct = round(maint * 100 / net_liq, 6) if (net_liq and maint and net_liq != 0) else 0
    cursor.execute(
        """
        INSERT INTO Capitals
            (AccountId, Date, NetLiquidity, Maintenance, ExcessLiquidity, Bpr,
             MaintenancePct, TotalPnl, UnrealizedPnl, RealizedPnl, NetDelta,
             NetTheta, NetVega, NetGamma, AvgIv, TotalMargin, TotalCommissions)
        VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
        """,
        account_id, dt, net_liq or 0, maint or 0, excess or 0, bpr or 0, maint_pct,
    )
    return 1


def insert_portfolio_if_missing(cursor: pyodbc.Cursor, account_id: int, row: tuple) -> int:
    budget = enum_val(row[0] if len(row) > 0 else None, BUDGET)
    strategy = enum_val(row[1] if len(row) > 1 else None, STRATEGY)
    if budget is None or strategy is None:
        return 0
    cursor.execute(
        "SELECT 1 FROM Portfolios WHERE AccountId = ? AND Budget = ? AND Strategy = ?",
        account_id, budget, strategy,
    )
    if cursor.fetchone():
        return 0
    min_alloc = num_val(row[2] if len(row) > 2 else None) or 0
    max_alloc = num_val(row[3] if len(row) > 3 else None) or 0
    cursor.execute(
        """
        INSERT INTO Portfolios (AccountId, Budget, Strategy, MinAllocation, MaxAllocation)
        VALUES (?, ?, ?, ?, ?)
        """,
        account_id, budget, strategy, min_alloc, max_alloc,
    )
    return 1


def insert_option_position_if_missing(cursor: pyodbc.Cursor, account_id: int, row: tuple) -> int:
    symbol = str_val(row[0] if len(row) > 0 else None)
    contract_id = str_val(row[1] if len(row) > 1 else None)
    if not symbol or not contract_id:
        return 0
    cursor.execute(
        "SELECT 1 FROM OptionPositions WHERE AccountId = ? AND ContractId = ?",
        account_id, contract_id,
    )
    if cursor.fetchone():
        return 0
    opened = date_val(row[2] if len(row) > 2 else None)
    expiry = date_val(row[3] if len(row) > 3 else None)
    closed = date_val(row[4] if len(row) > 4 else None)
    pos = int_val(row[5] if len(row) > 5 else None) or 0
    right = enum_val(row[6] if len(row) > 6 else None, POSITION_RIGHT, default=0) or 0
    strike = num_val(row[7] if len(row) > 7 else None) or 0
    cost = num_val(row[8] if len(row) > 8 else None) or 0
    commission = num_val(row[9] if len(row) > 9 else None) or 0
    multiplier = int_val(row[10] if len(row) > 10 else None) or 100
    try:
        con_id_int = int(contract_id)
    except ValueError:
        con_id_int = None
    cursor.execute(
        """
        INSERT INTO OptionPositions
            (AccountId, Symbol, ContractId, ConId, Opened, Expiry, Closed,
             Pos, [Right], Strike, Cost, Commission, Multiplier)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        account_id, symbol, contract_id, con_id_int,
        opened, expiry, closed, pos, right, strike, cost, commission, multiplier,
    )
    return 1


def insert_log_if_missing(cursor: pyodbc.Cursor, account_id: int, row: tuple,
                          existing_contract_ids: set[str]) -> int:
    """Insert a greeks log row only when the ContractId is already in OptionPositions."""
    dt = date_val(row[0] if len(row) > 0 else None)
    contract_id = str_val(row[1] if len(row) > 1 else None)
    if dt is None or not contract_id:
        return 0
    if contract_id not in existing_contract_ids:
        return 0
    cursor.execute(
        "SELECT 1 FROM OptionPositionsLogs WHERE AccountId = ? AND ContractId = ? AND DateTime = ?",
        account_id, contract_id, dt,
    )
    if cursor.fetchone():
        return 0
    def g(i: int) -> Any:
        return row[i] if i < len(row) else None

    cursor.execute(
        """
        INSERT INTO OptionPositionsLogs
            (AccountId, DateTime, ContractId, Underlying, Iv, Price, TimeValue,
             Delta, Theta, Gamma, Vega, Margin)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        account_id, dt, contract_id,
        safe_decimal(g(2)),
        safe_decimal(g(3)),
        safe_decimal(g(4)),
        safe_decimal(g(5)),
        safe_decimal(g(6)),
        safe_decimal(g(7)),
        safe_decimal(g(8)),
        safe_decimal(g(9)),
        safe_decimal(g(10)),
    )
    return 1


# ── Linker ─────────────────────────────────────────────────────────────

def link_trade_to_positions(
    cursor: pyodbc.Cursor,
    account_id: int,
    trade_id: int,
    trade: TradeRow,
    interactive: bool,
) -> tuple[int, int]:
    """Finds unassigned positions within ±1 day of trade.date and links them.
    Returns (options_linked, stocks_linked). Prompts user when nothing matches."""
    day_low = trade.date - timedelta(days=1)
    day_high = trade.date + timedelta(days=1)

    cursor.execute(
        """
        SELECT Id FROM OptionPositions
        WHERE AccountId = ? AND Symbol = ?
          AND Opened BETWEEN ? AND ?
          AND TradeId IS NULL
        """,
        account_id, trade.symbol, day_low, day_high,
    )
    option_ids = [int(r[0]) for r in cursor.fetchall()]

    cursor.execute(
        """
        SELECT Id FROM StockPositions
        WHERE AccountId = ? AND Symbol = ?
          AND Date BETWEEN ? AND ?
          AND TradeId IS NULL
        """,
        account_id, trade.symbol, day_low, day_high,
    )
    stock_ids = [int(r[0]) for r in cursor.fetchall()]

    if option_ids or stock_ids:
        if option_ids:
            cursor.execute(
                f"UPDATE OptionPositions SET TradeId = ? WHERE Id IN ({','.join('?' * len(option_ids))})",
                trade_id, *option_ids,
            )
        if stock_ids:
            cursor.execute(
                f"UPDATE StockPositions SET TradeId = ? WHERE Id IN ({','.join('?' * len(stock_ids))})",
                trade_id, *stock_ids,
            )
        return len(option_ids), len(stock_ids)

    # No auto-match — ask the user
    print(f"  NO MATCH for trade {trade_id}: {trade.symbol} {trade.date:%Y-%m-%d} "
          f"strat={trade.strategy} type={trade.type_of_trade}")
    if not interactive:
        return 0, 0

    # Show nearby unassigned positions for context
    cursor.execute(
        """
        SELECT Id, Opened, Strike, [Right], Pos FROM OptionPositions
        WHERE AccountId = ? AND Symbol = ? AND TradeId IS NULL
        ORDER BY ABS(DATEDIFF(day, Opened, ?))
        """,
        account_id, trade.symbol, trade.date,
    )
    near_opts = cursor.fetchall()
    cursor.execute(
        """
        SELECT TOP 5 Id, Date, PosChange, Price FROM StockPositions
        WHERE AccountId = ? AND Symbol = ? AND TradeId IS NULL
        ORDER BY ABS(DATEDIFF(day, Date, ?))
        """,
        account_id, trade.symbol, trade.date,
    )
    near_stocks = cursor.fetchall()

    if near_opts:
        print("    Nearest unassigned OptionPositions:")
        for r in near_opts[:5]:
            print(f"      opt#{r[0]} opened={r[1]:%Y-%m-%d} strike={r[2]} right={r[3]} pos={r[4]}")
    if near_stocks:
        print("    Nearest unassigned StockPositions:")
        for r in near_stocks:
            print(f"      stk#{r[0]} date={r[1]:%Y-%m-%d} qty={r[2]} price={r[3]}")

    answer = input("    Enter IDs to link (e.g. 'opt:12,14 stk:7' or 'skip'): ").strip()
    if not answer or answer.lower() == "skip":
        return 0, 0

    opts_to_link: list[int] = []
    stocks_to_link: list[int] = []
    for chunk in answer.replace(",", " ").split():
        if chunk.startswith("opt:"):
            opts_to_link.append(int(chunk[4:]))
        elif chunk.startswith("stk:"):
            stocks_to_link.append(int(chunk[4:]))

    if opts_to_link:
        cursor.execute(
            f"UPDATE OptionPositions SET TradeId = ? WHERE Id IN ({','.join('?' * len(opts_to_link))})",
            trade_id, *opts_to_link,
        )
    if stocks_to_link:
        cursor.execute(
            f"UPDATE StockPositions SET TradeId = ? WHERE Id IN ({','.join('?' * len(stocks_to_link))})",
            trade_id, *stocks_to_link,
        )
    return len(opts_to_link), len(stocks_to_link)


# ── Main orchestrator ──────────────────────────────────────────────────

def process_file(
    cursor: pyodbc.Cursor,
    account_id: int,
    path: Path,
    interactive: bool,
) -> dict[str, int]:
    print(f"\n== {path.name} ==")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    counts = {
        "trades_inserted": 0, "trades_linked_auto": 0, "trades_linked_manual": 0,
        "trades_unmatched": 0, "weekly_prep": 0, "capital": 0, "portfolio": 0,
        "option_positions": 0, "logs": 0,
    }

    new_trade_ids: list[tuple[int, TradeRow]] = []

    # --- Trades ---
    if "TradeLog" in wb.sheetnames:
        ws = wb["TradeLog"]
        is_2026 = "Portfolio" in wb.sheetnames
        trades = extract_trades_2026(ws) if is_2026 else extract_trades_2025(ws)
        for t in trades:
            new_id = insert_trade_if_missing(cursor, account_id, t)
            if new_id is not None:
                counts["trades_inserted"] += 1
                new_trade_ids.append((new_id, t))
        print(f"  Trades: {counts['trades_inserted']} inserted")

    # --- WeeklyPrep ---
    if "WeeklyPrep" in wb.sheetnames:
        for r in wb["WeeklyPrep"].iter_rows(min_row=2, values_only=True):
            counts["weekly_prep"] += insert_weeklyprep_if_missing(cursor, account_id, r)
        print(f"  WeeklyPreps: {counts['weekly_prep']} inserted")

    # --- Capital ---
    if "Capital" in wb.sheetnames:
        for r in wb["Capital"].iter_rows(min_row=2, values_only=True):
            counts["capital"] += insert_capital_if_missing(cursor, account_id, r)
        print(f"  Capitals: {counts['capital']} inserted")

    # --- Portfolio (2026 only) ---
    if "Portfolio" in wb.sheetnames:
        for r in wb["Portfolio"].iter_rows(min_row=2, values_only=True):
            counts["portfolio"] += insert_portfolio_if_missing(cursor, account_id, r)
        print(f"  Portfolios: {counts['portfolio']} inserted")

    # --- OptionPositions (2026 only) ---
    if "OptionPositions" in wb.sheetnames:
        for r in wb["OptionPositions"].iter_rows(min_row=2, values_only=True):
            counts["option_positions"] += insert_option_position_if_missing(cursor, account_id, r)
        print(f"  OptionPositions: {counts['option_positions']} inserted")

    # --- OptionPositionsLogs / PositionsLog ---
    log_sheet = "OptionPositionsLog" if "OptionPositionsLog" in wb.sheetnames else (
        "PositionsLog" if "PositionsLog" in wb.sheetnames else None
    )
    if log_sheet:
        cursor.execute("SELECT ContractId FROM OptionPositions WHERE AccountId = ?", account_id)
        existing_cids = {str(r[0]) for r in cursor.fetchall()}
        for r in wb[log_sheet].iter_rows(min_row=2, values_only=True):
            counts["logs"] += insert_log_if_missing(cursor, account_id, r, existing_cids)
        print(f"  Logs ({log_sheet}): {counts['logs']} inserted")

    wb.close()

    # --- Link new trades to positions (after positions are inserted) ---
    print(f"\n  Linking {len(new_trade_ids)} new trades to positions ...")
    for trade_id, trade in new_trade_ids:
        opts, stocks = link_trade_to_positions(cursor, account_id, trade_id, trade, interactive)
        if opts or stocks:
            counts["trades_linked_auto"] += 1
        else:
            counts["trades_unmatched"] += 1
    print(f"  Linked: {counts['trades_linked_auto']}, unmatched: {counts['trades_unmatched']}")

    return counts


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge Portfolio-Tracking xlsx into tradelog DB")
    parser.add_argument("--server", required=True)
    parser.add_argument("--database", required=True)
    parser.add_argument("--user", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument("--ibkr-account-id", required=True)
    parser.add_argument("--xlsx", action="append", required=True, help="Can be repeated")
    parser.add_argument("--non-interactive", action="store_true",
                        help="Skip prompts for unmatched trades")
    args = parser.parse_args()

    print(f"Connecting to {args.server}/{args.database} as {args.user} ...")
    conn = connect(args.server, args.database, args.user, args.password)
    cursor = conn.cursor()

    try:
        account_id = resolve_account_id(cursor, args.ibkr_account_id)
        print(f"Resolved AccountId={account_id} for IbkrAccountId={args.ibkr_account_id}")

        totals: dict[str, int] = {}
        for path_str in args.xlsx:
            path = Path(path_str)
            if not path.exists():
                print(f"ERROR: file not found: {path}", file=sys.stderr)
                conn.rollback()
                return 1
            counts = process_file(cursor, account_id, path, not args.non_interactive)
            for k, v in counts.items():
                totals[k] = totals.get(k, 0) + v

        print("\n== Totals ==")
        for k in sorted(totals):
            print(f"  {k}: {totals[k]}")

        answer = input("\nCommit these changes? (yes/no): ").strip().lower()
        if answer == "yes":
            conn.commit()
            print("Committed.")
        else:
            conn.rollback()
            print("Rolled back.")

    except Exception as e:
        conn.rollback()
        print(f"ERROR: {e}", file=sys.stderr)
        raise
    finally:
        conn.close()

    return 0


if __name__ == "__main__":
    sys.exit(main())
